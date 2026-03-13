"""
Geocodifica los clientes de Contactos 2.0.xlsx usando Google Geocoding API o HERE API.

Configuración:
  - Elige API: 'google' o 'here'
  - Pon tu API key abajo

Uso: python geocode_clientes.py
"""

import time
import csv
import requests
import openpyxl

# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN — Elige una API y pon tu key
# ══════════════════════════════════════════════════════════════
API = 'google'  # 'google' o 'here'

GOOGLE_API_KEY = 'AIzaSyDk2Vb73qDa_Hueqe-CelNR70u9s2oZChM'
HERE_API_KEY   = 'TU_API_KEY_DE_HERE'

# ══════════════════════════════════════════════════════════════

EXCEL_FILE = 'Contactos 2.0.xlsx'
CSV_OUTPUT = 'clientes_geocoded.csv'
SHEET_NAME = 'Clientes_Planta'
DELAY = 0.05  # 50ms entre peticiones (ambas APIs aguantan alto volumen)


def geocode_google(direccion, localidad, pais):
    """Geocodifica con Google Geocoding API."""
    query = f"{direccion}, {localidad}, {pais}"
    try:
        r = requests.get('https://maps.googleapis.com/maps/api/geocode/json', params={
            'address': query,
            'key': GOOGLE_API_KEY,
        }, timeout=10)
        data = r.json()
        if data['status'] == 'OK' and data['results']:
            loc = data['results'][0]['geometry']['location']
            return loc['lat'], loc['lng']
        # Fallback: solo localidad + país
        r = requests.get('https://maps.googleapis.com/maps/api/geocode/json', params={
            'address': f"{localidad}, {pais}",
            'key': GOOGLE_API_KEY,
        }, timeout=10)
        data = r.json()
        if data['status'] == 'OK' and data['results']:
            loc = data['results'][0]['geometry']['location']
            return loc['lat'], loc['lng']
    except Exception as e:
        print(f"  Error: {e}")
    return None, None


def geocode_here(direccion, localidad, pais):
    """Geocodifica con HERE Geocoding API."""
    query = f"{direccion}, {localidad}, {pais}"
    try:
        r = requests.get('https://geocode.search.hereapi.com/v1/geocode', params={
            'q': query,
            'apiKey': HERE_API_KEY,
        }, timeout=10)
        data = r.json()
        if data.get('items'):
            pos = data['items'][0]['position']
            return pos['lat'], pos['lng']
        # Fallback: solo localidad + país
        r = requests.get('https://geocode.search.hereapi.com/v1/geocode', params={
            'q': f"{localidad}, {pais}",
            'apiKey': HERE_API_KEY,
        }, timeout=10)
        data = r.json()
        if data.get('items'):
            pos = data['items'][0]['position']
            return pos['lat'], pos['lng']
    except Exception as e:
        print(f"  Error: {e}")
    return None, None


def geocodificar(direccion, localidad, pais):
    if API == 'google':
        return geocode_google(direccion, localidad, pais)
    else:
        return geocode_here(direccion, localidad, pais)


def main():
    print(f"API seleccionada: {API.upper()}")
    print(f"Leyendo {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    col_nombre = 0
    col_dir = 2
    col_loc = 3
    col_pais = 4
    col_lat = 5
    col_lon = 6

    clientes = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        nombre = row[col_nombre].value
        direccion = row[col_dir].value or ''
        localidad = row[col_loc].value or ''
        pais = row[col_pais].value or ''
        if nombre:
            clientes.append({
                'row': row[0].row,
                'nombre': nombre,
                'direccion': direccion,
                'localidad': localidad,
                'pais': pais,
            })

    total = len(clientes)
    print(f"Total clientes: {total}")
    print()

    csvfile = open(CSV_OUTPUT, 'w', newline='', encoding='utf-8')
    writer = csv.writer(csvfile)
    writer.writerow(['Nombre', 'Direccion', 'Localidad', 'Pais', 'Latitud', 'Longitud', 'OK'])

    ok_count = 0
    fail_count = 0

    for i, c in enumerate(clientes):
        print(f"[{i+1}/{total}] {c['nombre'][:50]}...", end=' ')

        lat, lon = geocodificar(c['direccion'], c['localidad'], c['pais'])

        if lat is not None:
            print(f"OK ({lat:.6f}, {lon:.6f})")
            ok_count += 1
            ws.cell(row=c['row'], column=col_lat + 1, value=lat)
            ws.cell(row=c['row'], column=col_lon + 1, value=lon)
            writer.writerow([c['nombre'], c['direccion'], c['localidad'], c['pais'], lat, lon, 'SI'])
        else:
            print("FALLO")
            fail_count += 1
            writer.writerow([c['nombre'], c['direccion'], c['localidad'], c['pais'], '', '', 'NO'])

        csvfile.flush()

        if (i + 1) % 100 == 0:
            print(f"  >>> Guardando progreso ({ok_count} OK, {fail_count} fallos)...")
            wb.save(EXCEL_FILE)

        time.sleep(DELAY)

    wb.save(EXCEL_FILE)
    csvfile.close()

    print()
    print(f"=== COMPLETADO ===")
    print(f"Geocodificados: {ok_count}/{total}")
    print(f"Fallidos: {fail_count}/{total}")
    print(f"CSV: {CSV_OUTPUT}")
    print(f"Excel actualizado: {EXCEL_FILE}")


if __name__ == '__main__':
    main()
