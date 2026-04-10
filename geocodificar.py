#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Geocodifica clientes del archivo 'Datos Velneo.xlsx' usando Nominatim (OpenStreetMap).
Agrupa por CP+Localidad+Provincia para minimizar peticiones.
Genera 'Datos Velneo Geocodificado.csv' con las coordenadas rellenadas.
"""

import openpyxl
import urllib.request
import urllib.parse
import json
import csv
import time
import os
import sys

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos Velneo.xlsx")
CSV_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datos Velneo Geocodificado.csv")
CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "geocode_cache.json")

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
HEADERS = {"User-Agent": "GestorRutas-Geocoder/1.0 (geocoding client addresses)"}
RATE_LIMIT = 1.1  # seconds between requests


def clean_province(prov):
    """'33 - ASTURIAS' -> 'ASTURIAS'"""
    if not prov:
        return ""
    if " - " in str(prov):
        return str(prov).split(" - ", 1)[1].strip()
    return str(prov).strip()


def clean_country(country):
    """'1 - ESPANA' -> 'Spain'"""
    if not country:
        return ""
    c = str(country)
    if " - " in c:
        c = c.split(" - ", 1)[1].strip()
    mapping = {
        "ESPANA": "Spain", "ESPAÑA": "Spain",
        "PORTUGAL": "Portugal",
        "ITALIA": "Italy",
        "FRANCIA": "France",
        "ALEMANIA": "Germany",
        "BELGICA": "Belgium", "BÉLGICA": "Belgium",
        "HOLANDA": "Netherlands", "PAISES BAJOS": "Netherlands",
        "COLOMBIA": "Colombia",
        "ECUADOR": "Ecuador",
        "KENIA": "Kenya", "KENYA": "Kenya",
        "ETIOPIA": "Ethiopia", "ETIOPÍA": "Ethiopia",
    }
    return mapping.get(c.upper(), c)


def clean_address(addr):
    """Remove _x000D_ and other artifacts"""
    if not addr:
        return ""
    s = str(addr)
    s = s.replace("_x000D_", "").replace("\r", "").replace("\n", " ").strip()
    # Remove trailing commas and whitespace
    s = s.strip(", ")
    return s


def geocode(query_parts, retries=2):
    """Geocode using Nominatim free-form or structured query."""
    # Try structured query first
    params = {}
    if query_parts.get("postalcode"):
        params["postalcode"] = query_parts["postalcode"]
    if query_parts.get("city"):
        params["city"] = query_parts["city"]
    if query_parts.get("state"):
        params["state"] = query_parts["state"]
    if query_parts.get("country"):
        params["country"] = query_parts["country"]

    params["format"] = "json"
    params["limit"] = "1"
    params["addressdetails"] = "0"

    url = NOMINATIM_URL + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers=HEADERS)

    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                if data:
                    return float(data[0]["lat"]), float(data[0]["lon"])
            break
        except Exception as e:
            if attempt < retries:
                time.sleep(2)
            else:
                print(f"  Error geocoding {params}: {e}", file=sys.stderr)
    return None, None


def geocode_freeform(query, retries=2):
    """Fallback: free-form geocoding."""
    params = {"q": query, "format": "json", "limit": "1"}
    url = NOMINATIM_URL + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers=HEADERS)

    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode("utf-8"))
                if data:
                    return float(data[0]["lat"]), float(data[0]["lon"])
            break
        except Exception as e:
            if attempt < retries:
                time.sleep(2)
    return None, None


def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=1)


def main():
    print("Leyendo archivo Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Hoja1"]

    headers = [cell.value for cell in ws[1]]
    print(f"  Columnas: {len(headers)}, Filas: {ws.max_row - 1}")

    # Read all rows
    rows_data = []
    unique_locations = {}  # key -> list of row indices

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
        row_list = list(row)
        rows_data.append(row_list)

        lat = row[13]
        lon = row[14]

        # Skip if already has valid coords
        if lat and lon and lat != 0 and lon != 0:
            continue

        cp = str(row[11]).strip() if row[11] else ""
        localidad = clean_address(str(row[12])) if row[12] else ""
        provincia = clean_province(row[8])
        pais = clean_country(row[7])

        key = f"{cp}|{localidad}|{provincia}|{pais}"
        if key not in unique_locations:
            unique_locations[key] = {
                "cp": cp, "localidad": localidad,
                "provincia": provincia, "pais": pais,
                "rows": []
            }
        unique_locations[key]["rows"].append(idx)

    print(f"  Ubicaciones unicas a geocodificar: {len(unique_locations)}")

    # Load cache
    cache = load_cache()
    cached_count = sum(1 for k in unique_locations if k in cache)
    print(f"  En cache: {cached_count}/{len(unique_locations)}")

    # Geocode unique locations
    total = len(unique_locations)
    done = 0
    found = 0
    not_found = 0
    from_cache = 0

    for key, loc in unique_locations.items():
        done += 1

        # Check cache
        if key in cache:
            lat, lon = cache[key]
            from_cache += 1
        else:
            # Build query
            cp = loc["cp"]
            city = loc["localidad"]
            state = loc["provincia"]
            country = loc["pais"]

            lat, lon = None, None

            # Strategy 1: Structured query with CP + city + state + country
            if cp or city:
                query_parts = {}
                if cp and cp != "None" and cp != "0":
                    query_parts["postalcode"] = cp
                if city:
                    query_parts["city"] = city
                if state:
                    query_parts["state"] = state
                if country:
                    query_parts["country"] = country

                if query_parts:
                    lat, lon = geocode(query_parts)
                    time.sleep(RATE_LIMIT)

            # Strategy 2: Free-form fallback with city + CP + country
            if lat is None and (city or cp):
                parts = [p for p in [city, cp, state, country] if p and p != "None" and p != "0"]
                if parts:
                    freeform = ", ".join(parts)
                    lat, lon = geocode_freeform(freeform)
                    time.sleep(RATE_LIMIT)

            # Strategy 3: Just city + country
            if lat is None and city and country:
                lat, lon = geocode_freeform(f"{city}, {country}")
                time.sleep(RATE_LIMIT)

            cache[key] = [lat, lon]

            # Save cache periodically
            if done % 50 == 0:
                save_cache(cache)

        if lat is not None:
            found += 1
            # Apply to all rows with this location
            for row_idx in loc["rows"]:
                rows_data[row_idx][13] = lat  # GPS Latitud
                rows_data[row_idx][14] = lon  # GPS longitud
        else:
            not_found += 1

        if done % 100 == 0 or done == total:
            print(f"  Progreso: {done}/{total} | Encontrados: {found} | No encontrados: {not_found} | Cache: {from_cache}")
            sys.stdout.flush()

    # Final cache save
    save_cache(cache)

    print(f"\nResultados finales:")
    print(f"  Total ubicaciones: {total}")
    print(f"  Geocodificadas: {found}")
    print(f"  No encontradas: {not_found}")
    print(f"  Desde cache: {from_cache}")

    # Write CSV output
    print(f"\nEscribiendo CSV: {CSV_OUTPUT}")

    # Clean headers for CSV
    clean_headers = []
    for h in headers:
        if h:
            clean_headers.append(str(h))
        else:
            clean_headers.append("")

    with open(CSV_OUTPUT, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(clean_headers)
        for row in rows_data:
            # Clean _x000D_ from all string fields
            cleaned = []
            for val in row:
                if isinstance(val, str):
                    cleaned.append(val.replace("_x000D_", "").replace("\r", ""))
                elif val is None:
                    cleaned.append("")
                else:
                    cleaned.append(val)
            writer.writerow(cleaned)

    print(f"Completado! Archivo generado: {CSV_OUTPUT}")

    # Count how many rows now have coordinates
    rows_with_coords = sum(1 for r in rows_data if r[13] and r[13] != 0)
    print(f"Filas con coordenadas: {rows_with_coords}/{len(rows_data)}")


if __name__ == "__main__":
    main()
