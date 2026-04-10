#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Cruza el Excel 'Datos Velneo.xlsx' con los clientes de la BD (db_clients_coords.csv)
para rellenar coordenadas GPS usando matching por nombre y dirección.
"""

import openpyxl
import csv
import re
import unicodedata
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "Datos Velneo.xlsx")
DB_CSV = os.path.join(BASE_DIR, "db_clients_coords.csv")
OUTPUT_CSV = os.path.join(BASE_DIR, "Datos Velneo Geocodificado.csv")


def normalize(text):
    """Normalize text for fuzzy comparison: lowercase, remove accents, strip punctuation."""
    if not text:
        return ""
    s = str(text).upper().strip()
    # Remove _x000D_ artifacts
    s = s.replace("_X000D_", "").replace("_x000D_", "").replace("\r", "").replace("\n", " ")
    # Normalize unicode
    s = unicodedata.normalize("NFKD", s)
    # Remove accents
    s = "".join(c for c in s if not unicodedata.combining(c))
    # Lowercase
    s = s.lower()
    # Remove common suffixes/prefixes
    s = re.sub(r'\b(s\.?l\.?|s\.?a\.?|s\.?l\.?u\.?|soc\.?\s*coop\.?)\b', '', s)
    # Remove punctuation except spaces
    s = re.sub(r'[^a-z0-9\s]', ' ', s)
    # Collapse whitespace
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_business_name(excel_desc):
    """Extract business name from 'BUSINESS NAME / PERSON NAME' format."""
    if not excel_desc:
        return ""
    s = str(excel_desc)
    # Split on ' / ' to get business name (first part)
    if " / " in s:
        s = s.split(" / ")[0].strip()
    return s


def build_address_key(address):
    """Build a normalized address key."""
    if not address:
        return ""
    s = normalize(address)
    # Remove common address prefixes
    s = re.sub(r'\b(c/|c/\.|calle|avda|av\.|avenida|rua|plaza|pza|ctra|carretera|paseo)\b', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def similarity_score(s1, s2):
    """Simple word-overlap similarity between two normalized strings."""
    if not s1 or not s2:
        return 0
    words1 = set(s1.split())
    words2 = set(s2.split())
    if not words1 or not words2:
        return 0
    intersection = words1 & words2
    # Jaccard-like but weighted by shorter set
    shorter = min(len(words1), len(words2))
    if shorter == 0:
        return 0
    return len(intersection) / shorter


def main():
    # 1. Load DB clients with coords
    print("Cargando clientes de la BD...")
    db_clients = []
    db_name_index = {}  # normalized name -> list of db clients
    db_addr_index = {}  # normalized addr key -> list of db clients

    with open(DB_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            try:
                x = float(row["x"]) if row["x"] else 0
                y = float(row["y"]) if row["y"] else 0
            except ValueError:
                continue

            client = {
                "id": row["id"],
                "name": row["name"],
                "name_norm": normalize(row["name"]),
                "address": row["address"],
                "addr_norm": build_address_key(row["address"]),
                "postcode": row.get("postcode", ""),
                "x": x,
                "y": y,
            }
            db_clients.append(client)

            # Index by normalized name words
            nname = client["name_norm"]
            if nname:
                # Index by each significant word (3+ chars)
                for word in nname.split():
                    if len(word) >= 3:
                        db_name_index.setdefault(word, []).append(client)

    print(f"  Cargados: {len(db_clients)} clientes con coordenadas")

    # 2. Load Excel
    print("Cargando Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Hoja1"]
    headers = [cell.value for cell in ws[1]]

    rows_data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows_data.append(list(row))

    print(f"  Filas Excel: {len(rows_data)}")

    # 3. Match each Excel row to DB
    print("\nCruzando datos...")
    matched = 0
    not_matched = 0
    already_has = 0
    match_details = {"exact": 0, "business_name": 0, "fuzzy": 0, "address": 0}

    for idx, row in enumerate(rows_data):
        lat = row[13]
        lon = row[14]

        # Skip if already has valid coordinates
        if lat and lon and lat != 0 and lon != 0:
            already_has += 1
            continue

        excel_desc = str(row[1]) if row[1] else ""
        business_name = extract_business_name(excel_desc)
        excel_addr = str(row[9]) if row[9] else ""
        excel_cp = str(row[11]).strip() if row[11] else ""
        excel_loc = str(row[12]).strip() if row[12] else ""

        norm_desc = normalize(excel_desc)
        norm_business = normalize(business_name)
        norm_addr = build_address_key(excel_addr)

        best_match = None
        best_score = 0
        match_type = ""

        # Strategy 1: Exact normalized name match
        for db_c in db_clients:
            if db_c["name_norm"] == norm_desc or db_c["name_norm"] == norm_business:
                best_match = db_c
                best_score = 1.0
                match_type = "exact"
                break

        # Strategy 2: Check if DB name is contained in Excel name or vice versa
        if not best_match and norm_business:
            for db_c in db_clients:
                db_name = db_c["name_norm"]
                if not db_name:
                    continue
                if db_name in norm_business or norm_business in db_name:
                    score = min(len(db_name), len(norm_business)) / max(len(db_name), len(norm_business))
                    if score > best_score and score >= 0.5:
                        best_match = db_c
                        best_score = score
                        match_type = "business_name"

        # Strategy 3: Word-overlap fuzzy matching using index
        if not best_match and norm_business:
            candidates = set()
            for word in norm_business.split():
                if len(word) >= 3 and word in db_name_index:
                    for c in db_name_index[word]:
                        candidates.add(id(c))

            for db_c in db_clients:
                if id(db_c) not in candidates:
                    continue
                score = similarity_score(norm_business, db_c["name_norm"])
                if score > best_score and score >= 0.6:
                    best_match = db_c
                    best_score = score
                    match_type = "fuzzy"

        # Strategy 4: Address + CP match (if name didn't work)
        if not best_match and norm_addr and excel_cp:
            for db_c in db_clients:
                if db_c["postcode"] and db_c["postcode"] == excel_cp:
                    addr_score = similarity_score(norm_addr, db_c["addr_norm"])
                    if addr_score > best_score and addr_score >= 0.5:
                        best_match = db_c
                        best_score = addr_score
                        match_type = "address"

        if best_match:
            rows_data[idx][13] = best_match["x"]
            rows_data[idx][14] = best_match["y"]
            matched += 1
            match_details[match_type] += 1
        else:
            not_matched += 1

        if (idx + 1) % 1000 == 0:
            print(f"  Progreso: {idx+1}/{len(rows_data)} | Matched: {matched} | No match: {not_matched}")

    print(f"\n--- Resultados ---")
    print(f"Total filas:          {len(rows_data)}")
    print(f"Ya tenian coords:     {already_has}")
    print(f"Matched (cruzados):   {matched}")
    print(f"  - Exacto:           {match_details['exact']}")
    print(f"  - Nombre negocio:   {match_details['business_name']}")
    print(f"  - Fuzzy:            {match_details['fuzzy']}")
    print(f"  - Por direccion:    {match_details['address']}")
    print(f"Sin match:            {not_matched}")

    # 4. Write CSV output
    print(f"\nEscribiendo: {OUTPUT_CSV}")
    clean_headers = [str(h) if h else "" for h in headers]

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(clean_headers)
        for row in rows_data:
            cleaned = []
            for val in row:
                if isinstance(val, str):
                    cleaned.append(val.replace("_x000D_", "").replace("\r", ""))
                elif val is None:
                    cleaned.append("")
                else:
                    cleaned.append(val)
            writer.writerow(cleaned)

    # 5. Also write a report of unmatched rows
    report_path = os.path.join(BASE_DIR, "sin_coordenadas.csv")
    unmatched_count = 0
    with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Codigo", "Descripcion", "Direccion", "CP", "Localidad", "Provincia"])
        for row in rows_data:
            lat = row[13]
            lon = row[14]
            if not lat or not lon or lat == 0 or lon == 0:
                writer.writerow([row[0], row[1], row[9], row[11], row[12], row[8]])
                unmatched_count += 1

    print(f"Filas sin coordenadas exportadas a: {report_path} ({unmatched_count} filas)")
    print("Completado!")


if __name__ == "__main__":
    main()
